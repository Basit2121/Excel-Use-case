from openpyxl import load_workbook, Workbook
import random
import string

def generate_random_string():
    return ''.join(random.choices(string.ascii_letters, k=8))

def enter_data_in_sheet(filename):

    try:
        # Try to load the existing workbook
        wb = load_workbook(filename)
        ws = wb.active
    except FileNotFoundError:
        # If the file doesn't exist, create a new workbook
        wb = Workbook()
        ws = wb.active

    # Find the last row with data in column A
    max_row = ws.max_row

    for i in range(100):

        random_string = generate_random_string()
        ws[f'A{max_row+i+1}'] = f'{random_string}'
        random_string = generate_random_string()
        ws[f'B{max_row+i+1}'] = f'{random_string}'

    # Save the workbook
    wb.save(filename)

def check_filled(filename):

    wb = load_workbook(filename)
    ws = wb.active

    filled_row_count = 0

    for row in ws.iter_rows(values_only=True):

        filled_row_count = filled_row_count + 1

        for cell_value in row:
            if cell_value:
                print(f"Cell is filled with: {cell_value}")
    
    return filled_row_count

if __name__ == "__main__":
    #val = check_filled("example.xlsx")
    #print(val)
    enter_data_in_sheet("example.xlsx")