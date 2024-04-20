from openpyxl import load_workbook, Workbook

def create_account(filename, username, password):

    try:
        # Try to load the existing workbook
        wb = load_workbook(filename)
        ws = wb.active
    except FileNotFoundError:
        # If the file doesn't exist, create a new workbook
        wb = Workbook()
        ws = wb.active

    # Find the last row with data in column A
    max_row = ws.max_row
 
    ws[f'A{max_row+1}'] = username
    ws[f'B{max_row+1}'] = password

    # Save the workbook
    wb.save(filename)

def check_login(filename, username, password):

    wb = load_workbook(filename)
    ws = wb.active

    incorrect_user_or_pass = True

    for row in ws.iter_rows(values_only=True):
        
        if username == row[0] and password == row[1]:
            print("Logged In.")
            incorrect_user_or_pass = False
        
    if incorrect_user_or_pass == True:
        print("Username or Password is Incorrect.")

if __name__ == "__main__":

    choice = int(input("Enter 1 to Create Account\nEnter 2 to Log In to Account"))

    username = input("Username :")
    password = input("Password :")

    if choice == 1:
        create_account("example.xlsx", username, password)
    elif choice == 2:
        check_login("example.xlsx", username, password)#rwmGEKLX	CspQSPTP BANSwyht
